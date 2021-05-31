'''
If you're reading this, I'm so sorry
This was the first program I wrote for this company
Since then, I have learned a lot, but never had the chance to return and refactor
'''

''' TODO
-Test warnable clients
-Auto update PAPS/BoL according to Carrier Tracking.xlsx
-Set per-name 800$ limit on S321 manifests
-Throw error when 800$ limit is exceeded
-Add manifest emailing
-Check for duplicate SCNs in historic orders
-Add Last Updated to Master_FDA_LIST
-Change filewriting to not overwrite

Password
CONFIG not saving?
'''

import json
from datetime import date
import time
import os
import csv
import sys
import traceback
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from email.mime.base import MIMEBase
from email import encoders
import re
import shutil

#installable through PiP
import openpyxl as pyxl
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.utils import Image

#Optional library
#import Google_API_Tool as gAPI

#Optional library, included by /appJar/ folder
from appJar import gui

### Functions

## Mainline Functions

def doEverything():
	#Creates the core data used for every other function
	readMasterMetadata() #Merges XLSX report (carrier, client, tracking number, close date) with ACE report (shipping info) to create one unified data source
	constructACE() #Merges XLSX report (carrier, client, tracking number, close date) with ACE report (shipping info) to create one unified data source
	               #Then filters out all ACE entries that don't match a batch scan. Also outputs ACE
	assignGaylords() #Creates a list of all gaylords, which carrier (eg. DHL, USPS, FedEx) they belong to, if it has FDA-regulated products

	#Uses the core data to produce paperwork 
	createDetailedReport() #Th e report Tri-Ad uses to find packages when searched
	createLoadSheet() #Creates a useful but non-necessary info sheet
	createBoL() #Creates the Bill of lading for this load
	createIMSBoL() #Creates the Bill of Lading for IMS
	createProForma() #Creates the XLSX file we upload to SmartBorder to create our Proforma invoice
	#updateIMSTracker() #Disabled currently as the google apps API keeps giving misleading errors

	app.infoBox("Done", "All paperwork completed!")

def readMasterMetadata():
	global master_metadata
	global config_data
	master_metadata["date"] = app.getEntry("Date:")
	master_metadata["BoL"] = app.getEntry("BoL #:")
	master_metadata["PAPS"] = app.getEntry("PAPS #:")
	master_metadata["SCAC"] = config_data["SCAC"]
	app.setEntry("File Date:", master_metadata["date"])
	createOutputFolder(master_metadata["date"])

def createConsolidatedJSON():
	#Opens the ACE and XLSX report. Matches entries in the two based on order IDs. Takes the client/carrier/close date data from the XLSX and adds it to the ACE
	global master_ACE_data

	#Loads unprocessed ACE manifest
	try:
		ACE_data = []
		with open(app.getEntry("ACEManifestFileEntry"), "r") as ACE_file:
			ACE_data = json.load(ACE_file)
	except: errorBox("Error loading ACE Manifest.\nFile open in another program?")

	#Loads XLSX Report
	try:
		file_path = app.getEntry("XLSXReportFileEntry")
		worksheet = pyxl.load_workbook(file_path).active

		#Load Header
		XLSX_header = ()
		for row in worksheet.values: #Looks odd but openpyxl doesn't like to just _let_ you access the values of the first row
			XLSX_header = row
			break #Effectively sets XLSX_header <= row[0]

		global config_data
		ORDERID_column_index = -1
		client_name_column_index = -1
		carrier_column_index = -1
		ship_date_column_index = -1
		tracking_number_column_index = -1

		#Figures out where all the important columns are, in order to load that data
		#Can change the names of the important columns in CONFIG.json
		for i, cell in enumerate(XLSX_header):
			if XLSX_header[i] == config_data["XLSX_Report_ORDERID_column_name"]:
				ORDERID_column_index = i
			if XLSX_header[i] == config_data["XLSX_Report_client_name_column_name"]:
				client_name_column_index = i
			if XLSX_header[i] == config_data["XLSX_Report_carrier_column_name"]:
				carrier_column_index = i
			if XLSX_header[i] == config_data["XLSX_Report_ship_date_column_name"]:
				ship_date_column_index = i
			if XLSX_header[i] == config_data["XLSX_Report_tracking_number_column_name"]:
				tracking_number_column_index = i

		#Loads unfiltered Data
		XLSX_data = []
		for row in worksheet.values:
			XLSX_data.append(row)

		#Matches XLSX Report data to ACE Manifest data using ORDERID
		#Then adds the important data from the XLSX report to the ACE
		consolidated_ACE_data = []
		for XLSX_line in XLSX_data:
			for json_entry in ACE_data:
				if XLSX_line[ORDERID_column_index] == json_entry["ORDERID"]: #TODO
					json_entry["client"] = XLSX_line[client_name_column_index]
					json_entry["carrier"] = XLSX_line[carrier_column_index]
					json_entry["closeDate"] = XLSX_line[ship_date_column_index]
					json_entry["trackingNumber"] = XLSX_line[tracking_number_column_index]
					consolidated_ACE_data.append(json_entry)
		return consolidated_ACE_data
	except:	errorBox("Error loading XLSX data.\nFile open in another program?")

def loadFDASKUs():
	try:
		#NOTE: Currently if a client has FDA-regulated goods, none of their products can go through Section 321 (aka end up on the ACE)
		#This may possibly change in the future
		FDA_SKUs_list = []
		with open("resources/MASTER_FDA_LIST.csv", "r") as SKUs_file:
			csv_reader = csv.reader(SKUs_file, delimiter = ",")
			for line in csv_reader:
				if line != "":
					FDA_SKUs_list.append(line[2])
		FDA_SKUs_list.pop(0)
		return FDA_SKUs_list
	except:
		errorBox("Error loading resources/MASTER_FDA_LIST.csv.\nFile missing/open in another program?")

def loadBatchesFile():
	#Can load either the (old) Batches Scans or the (newly implemented) Detailed Report
	global config_data

	if app.getEntry("batchesFileEntry")[-4:] == ".csv":
		try:
			#Loads CSV Data
			csv_data = []
			with open(app.getEntry("batchesFileEntry"), "r") as batches_file:
				csv_reader = csv.reader(batches_file, delimiter = ",")
				for row in csv_reader:
					csv_data.append(row)

			#Reads the ehader to figure out which column is for Batches and which is for the gaylord its assigned to
			csv_header = csv_data[0]
			batch_index = -1
			gaylord_index = -1
			batches_data = []
			for i, cell in enumerate(csv_header):
				if cell == config_data["BATCHES_SCANS_batch_column_name"]:
					batch_index = i
				if cell == config_data["BATCHES_SCANS_gaylord_column_name"]:
					gaylord_index = i
			print("Batch Index:", batch_index, "Gaylord Index:" , gaylord_index)
			csv_data.pop(0) #Remove header from CSV data

			#Builds the list of batch-gaylord assignments
			for row in csv_data:
				batches_data.append({"batch": str(row[batch_index]), "gaylord": str(row[gaylord_index])})
			return batches_data
		except:
			errorBox("Error loading Batches Scans.\nFile open in another program?")

	elif app.getEntry("batchesFileEntry")[-5:] == ".xlsx": #If user is uploading Detailed Report
		try:

			file_path = app.getEntry("batchesFileEntry")
			workbook = pyxl.load_workbook(file_path)
			worksheet = workbook[config_data["Detailed_Report_scan_sheet_name"]]

			#Load Header from Detailed Report
			header = ()
			for row in worksheet.values:
				header = row
				break

			#Figure out where the batch column is and where the gaylord column is
			batch_index = -1
			gaylord_index = -1
			for i, cell in enumerate(header):
				if cell == config_data["Detailed_Report_batch_column_name"]:
					batch_index = i
				if cell == config_data["Detailed_Report_gaylord_column_name"]:
					gaylord_index = i

			#Loads raw XLSX values
			XLSX_data = []
			for row in worksheet.values:
				XLSX_data.append(row)
			XLSX_data.pop(0) # Removes header

			#Builds the list of batch-gaylord assignments
			batches_data = []
			for row in XLSX_data:
				if row[batch_index] != None and row[gaylord_index] != None:
					batches_data.append({"batch": str(row[batch_index]), "gaylord": str(row[gaylord_index])})

			return batches_data

		except:
			errorBox("Error loading Detailed Report.\nFile open in another program?")

def constructACE():
	global master_ACE_data #The outputted master ACE with all products
	global master_metadata #Details about the shipment (eg. date, BoL #, PAPS #, etc)
	
	batches_data = loadBatchesFile()
	FDA_clients_list = loadFDASKUs()

	#compare lists and build outputs
	master_ACE_data = [] 
	good_json = [] #The ACE that gets outputted, without FDA products
	good_batches = [] #Used to identify batches that did not match and ACE entries
	consolidated_json = createConsolidatedJSON() #Adds client/carrier/close date to ACE manifest for sorting purposes

	#The main loop that started this entire program
	#Matches Batches Scans to ACE entries
	#Only ACE entries that match a batch scan are added to the out-bound ACE
	for json_entry in consolidated_json:
		for row in batches_data:
			if json_entry["BATCHID"] == row["batch"] or json_entry["ORDERID"] == row["batch"]: #If there's a match
				if json_entry["commodities"][0]["description"] not in FDA_clients_list: #If the product is not commercially cleared
					json_entry["GAYLORD"] = row["gaylord"] #Append the Gaylord assignment to the entry
					good_json.append(json_entry) #Append to the ACE that goes to BorderConnect
					master_ACE_data.append(json_entry) #Append to the master ACE used for lots of stuff
					good_batches.append(row["batch"]) #Used to identify which batches aren't matched
					json_entry["shipmentClearance"] = "S321"
				else: #If it is commercially cleared
					json_entry["GAYLORD"] = row["gaylord"]
					#good_json.append(json_entry) #Doesn't go on the JSON that gets uploaded to BorderConnect, only Detailed Report (and the ProForma)
					master_ACE_data.append(json_entry)
					good_batches.append(row["batch"])
					json_entry["shipmentClearance"] = "FDA"
					
	good_json = validateJSON(good_json) #Removes duplicates and common errors that prevent ACE uploading

	#Error Recording
	with open(master_metadata["date"] + os.sep + master_metadata["date"] + "-Error_file.txt", "w") as error_file:

		#Warn unmatched JSON entries
		unmatched_json_list = []
		unmatched_json_uniques = []
		for json_entry in consolidated_json:
			if json_entry not in master_ACE_data and json_entry["BATCHID"] not in good_batches:
				if json_entry["BATCHID"] not in unmatched_json_uniques:
					unmatched_json_uniques.append(json_entry["BATCHID"])
					unmatched_json_list.append({"BATCHID": json_entry['BATCHID'], "closeDate": json_entry['closeDate'], "client": json_entry["client"]})
					#print("Unmatched JSON Entry: {} {} {}".format(json_entry['BATCHID'].ljust(10, " "), json_entry['closeDate'].ljust(22, " "), json_entry["client"]))
		#Send out an errorBox with unmatched JSON entries
		if unmatched_json_list != []:
			unmatched_json_list.sort(key = lambda x: x["closeDate"], reverse = True)
			print("\n>Printing Unmatched JSON Entries")
			print("If an order below is from today, it is likely a missing pick ticket that must be found and scanned")
			out_text = "Unmatched JSON Entries\nIf an order below is from today, it is likely a missing pick ticket that must be found and scanned\n"
			error_file.write("\n>Printing Unmatched JSON Entries")
			error_file.write("\nIf an order below is from today, it is likely a missing pick ticket that must be found and scanned")
			for line in unmatched_json_list:
				out_line = "Unmatched JSON Entry: {} {} {}".format(line['BATCHID'].ljust(10, " "), line['closeDate'].ljust(22, " "), line["client"])
				print(out_line)
				out_text += "\n" + out_line
				error_file.write("\n" + out_line)
			errorBox(out_text)

		#Warn unmatched Batch Scans
		unmatched_batch_list = []
		unmatched_batch_uniques = []
		for row in batches_data:
			if row["batch"] not in good_batches:
				if row["batch"] not in unmatched_batch_uniques:
					unmatched_batch_uniques.append(row["batch"])
					unmatched_batch_list.append({"batch": row["batch"], "gaylord": row["gaylord"]})
					#print("Unmatched batch: {} {}".format(row["batch"].ljust(8, " "), row["gaylord"]))
		#Send out an errorBox with unmatched batches
		if unmatched_batch_list != []:
			print("\n>Printing Unmatched Batches")
			print("Use the Batches page in Techship to find out what manfiest these belong to, and include them in the ACE")
			out_text = "Unmatched Batches\nUse the Batches page in Techship to find out what manfiest these belong to, and include them in the ACE"
			error_file.write("\n>Printing Unmatched Batches")
			error_file.write("\nUse the Batches page in Techship to find out what manfiest these belong to, and include them in the ACE")
			for line in unmatched_batch_list:
				out_line= "Unmatched batch: {} {}".format(line["batch"].ljust(8, " "), line["gaylord"])
				print(out_line)
				out_text += "\n" + out_line
				error_file.write("\n" + out_line)
			errorBox(out_text)

	#Package Count
	package_count = len(master_ACE_data)
	master_metadata["packageCount"] = package_count
	master_metadata["totalWeight"] = int(package_count // 2.20462) #Assumes 1 package ~= 1 lb, then converted to KG
	
	#Print ACE Manifest
	with open(master_metadata["date"] + os.sep + master_metadata["date"] + "-ACE.json", "w") as good_json_file:
		json.dump(good_json, good_json_file, indent = 4)

def validateJSON(in_json):
	#Removes duplicate entries from the ACE
	#Also automatically "corrects" a lot of errors that may prevent the ACE from being accepted by BorderConnect

	global config_data
	out_json = []

	# Get rid of duplicates
	for entry in in_json:

		do_once_flag = True
		if do_once_flag and entry["client"] in config_data["warnable_clients"]:
			app.errorBox("WARNING!", f"Warning! packages found for {entry['client']}")
			do_once_flag = False

		#LUS tickets can be split eg. "G1,2"
		#This method assigns the entry to the first gaylord (eg. G1) if it belongs to USPS and the second gaylord (eg. G2) otherwise
		if entry["client"] == "LUS Brands" and "," in entry["GAYLORD"]:
			if entry["carrier"] == "EHUB":
				entry["GAYLORD"] = entry["GAYLORD"].split(",")[0]
			else:
				entry["GAYLORD"] = "G" + entry["GAYLORD"].split(",")[1].replace("G", "")
		elif "," in entry["GAYLORD"]:
			app.errorBox("ErrorBox", f"WARNING! Non-LUS pick ticket assigned to multiple gaylords!\n{entry['BATCHID']} {entry['GAYLORD']} {entry['client']}")

		#Replaces SCAC (Carrier code) in SCN (Shipment Control Number)
		entry["shipmentControlNumber"] = entry["shipmentControlNumber"].replace("TAIW", app.getEntry("SCAC:"))
		
		#Override International Orders (set consignee to IMS)
		if entry["consignee"]["address"]["country"] != "US":
			entry["consignee"]["address"]["addressLine"] = config_data["consigneeOverrideAddress"]["addressLine"]
			entry["consignee"]["address"]["country"] = config_data["consigneeOverrideAddress"]["country"]
			entry["consignee"]["address"]["city"] = config_data["consigneeOverrideAddress"]["city"]
			entry["consignee"]["address"]["stateProvince"] = config_data["consigneeOverrideAddress"]["stateProvince"]
			entry["consignee"]["address"]["postalCode"] = config_data["consigneeOverrideAddress"]["postalCode"]

		#Validate Name
		if len(entry["consignee"]["name"]) <= 2:
			entry["consignee"]["name"] = entry["consignee"]["name"].ljust(3, "A")
		elif len(entry["consignee"]["name"]) >= 60:
			entry["consignee"]["name"] = entry["consignee"]["name"][:59]
		#Validate Address
		if len(entry["consignee"]["address"]["addressLine"]) <= 2:
			entry["consignee"]["address"]["addressLine"] = entry["consignee"]["address"]["addressLine"].ljust(3, "A")
		elif len(entry["consignee"]["address"]["addressLine"]) >= 55:
			entry["consignee"]["address"]["addressLine"] = entry["consignee"]["address"]["addressLine"][:54]
		#Validate City
		if len(entry["consignee"]["address"]["city"]) <= 3:
			entry["consignee"]["address"]["city"] = entry["consignee"]["address"]["city"].ljust(2, "A")
		elif len(entry["consignee"]["address"]["city"]) >= 30:
			entry["consignee"]["address"]["city"] = entry["consignee"]["address"]["city"][:29]
		#Validate State
		states_list = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "AS", "DC", "GU", "MP", "PR", "VI"]
		if entry["consignee"]["address"]["stateProvince"] not in states_list:
			entry["consignee"]["address"]["stateProvince"] = "NY"
		#Validate Zip
		if len(entry["consignee"]["address"]["postalCode"]) != 5:
			entry["consignee"]["address"]["postalCode"] = entry["consignee"]["address"]["postalCode"].rjust(5, "0")

		#Clean non-alphanumeric Characters
		entry["consignee"]["name"] = cleanString(entry["consignee"]["name"])
		entry["consignee"]["address"]["addressLine"] = cleanString(entry["consignee"]["address"]["addressLine"])
		entry["consignee"]["address"]["city"] = cleanString(entry["consignee"]["address"]["city"])

		#Check for duplicates
		if entry not in out_json:
			out_json.append(entry)
	return out_json

def assignGaylords():
	#Figures out which gaylords go where, how many total, and carrying what
	global master_metadata
	global master_ACE_data
	global config_data
	unique_gaylords_list = []
	master_metadata["gaylordAssignments"] = []

	for entry in master_ACE_data:
		#Clean it first
		if len(entry["GAYLORD"]) == 2:
			entry["GAYLORD"] = "G0" + entry["GAYLORD"][-1]
		#Add it to the uniques list
		match = False
		for unique_gaylord in unique_gaylords_list:
			if entry["GAYLORD"] == unique_gaylord["id"]:
				match = True
		if not match:
			unique_gaylords_list.append({"id": entry["GAYLORD"], "hasFDA": False, "hasUSPS": False, "hasDHL": False, "hasFedex": False, "packages": 0})
	
	#For every gaylord it checks the master file to see where the gaylord is heading and if it has FDA products
	FDA_clients_list = loadFDASKUs()
	for entry in master_ACE_data:
		for line in unique_gaylords_list:
			if entry["GAYLORD"] == line["id"]: #Will match only one gaylord
				#Now set flags
				line["packages"] += 1
				if entry["carrier"] in config_data["DHL_carriers_names"]:
					line["hasDHL"] = True
				elif entry["carrier"] in config_data["FEDEX_carriers_names"]:
					line["hasFedex"] = True
				elif entry["carrier"] in config_data["USPS_carriers_names"]:
					line["hasUSPS"]= True
				else:
					errorBox("Package with Order ID {} found without carrier".format(entry["ORDERID"]))
				if entry["shipmentClearance"] == "FDA":
					line["hasFDA"] = True

	#Checks if the gaylord has multiple carriers assigned and counts gaylords
	usps_count,	dhl_count, fedex_count = 0, 0, 0
	for line in unique_gaylords_list:
		out = {}
		out["id"] = line["id"]

		carriers_count = 0
		if line["hasUSPS"]: carriers_count += 1
		if line["hasDHL"]: carriers_count += 1
		if line["hasFedex"]: carriers_count += 1

		if carriers_count != 1:
			errorBox("Multiple carriers found for {}".format(line["id"]))
		elif line["hasUSPS"]:
			out["carrier"] = "EHUB"
			usps_count += 1
		elif line["hasDHL"]:
			out["carrier"] = "DHLGLOBALMAIL"
			dhl_count += 1
		elif line["hasFedex"]:
			out["carrier"] = "FEDEX"
			fedex_count += 1
			
		if line["hasFDA"]: out["hasFDA"] = "FDA"
		else: out["hasFDA"] = ""

		out["packageCount"] = line["packages"]
		master_metadata["gaylordAssignments"].append(out)

	master_metadata["USPSCount"] = usps_count
	master_metadata["DHLCount"] = dhl_count
	master_metadata["FEDEXCount"] = fedex_count
	master_metadata["gaylordCount"] = (usps_count + dhl_count + fedex_count)

	master_metadata["gaylordAssignments"].sort(key = lambda x: x["id"])

## Paperwork functions

def createDetailedReport():
	# The report used by Tri-Ad/us to figure out what packages belong to what gaylord/shipping information
	global master_ACE_data
	global master_metadata
	file_name = master_metadata["date"] + os.sep + master_metadata["date"] + "-Detailed_Report.csv"
	try:
		with open(file_name, "w", newline = "") as report_file:
			csv_writer = csv.writer(report_file)
			csv_writer.writerow(["Gaylord", "Name", "Address", "City", "State", "Country", "ZIP", "BATCHID", "ORDERID", "SCAC", "Service", "Client", "Close Date", "S321/FDA?", "Total Value", "Commodity 1", "Commodity 2", "Commodity 3", "Commodity 4", "Commodity 5", "Commodity 6", "Commodity 7", "Commodity 8", "Commodity 9", "Commodity 10"])
			for entry in master_ACE_data:
				#Ugly csv line building. Please ignore
				total_value = 0.0
				commodities_names_list = []
				for commodity in entry["commodities"]:
					commodities_names_list.append(commodity["description"])
					if commodity.get("value"):
						total_value += float(commodity.get("value"))
				out_line = [entry["GAYLORD"], entry["consignee"]["name"], entry["consignee"]["address"]["addressLine"], entry["consignee"]["address"]["city"], entry["consignee"]["address"]["stateProvince"], entry["consignee"]["address"]["country"], entry["consignee"]["address"]["postalCode"], entry["BATCHID"], entry["ORDERID"], entry["shipmentControlNumber"], entry["carrier"], entry["client"], entry["closeDate"], entry["shipmentClearance"], str(total_value)] + commodities_names_list
				csv_writer.writerow(out_line)
	except:
		errorBox(f"Error on line {entry['ORDERID']}.\nDetailed Report open in another program?")

def createBoL():
	#Loads a .jpg file template, then writes the last few details to the image before saving it as a .pdf
	global master_metadata
	file_name =  master_metadata["date"] + os.sep + master_metadata["date"] + "-Stalco-BoL.pdf"
	image_file_path = "resources/STALCO_BOL.jpg"
	c = canvas.Canvas(file_name, pagesize = (1668, 1986), bottomup = 0)
	image = ImageReader(image_file_path)
	c.drawImage(image, 0, 0, mask = "auto")

	c.setFont("Courier", 24)
	c.drawString(4, 340, app.getEntry("Date:"))
	c.drawString(1555, 330, master_metadata["BoL"])
	c.drawString(1562, 355, master_metadata["PAPS"])
	c.drawString(134, 1210, str(master_metadata["gaylordCount"]))
	c.drawString(176, 1310, str(master_metadata["USPSCount"]))
	c.drawString(171, 1360, str(master_metadata["DHLCount"]))
	c.drawString(181, 1410, str(master_metadata["FEDEXCount"]))
	c.drawString(525, 1175, str(master_metadata["packageCount"]))
	c.drawString(854, 1175, str(master_metadata["totalWeight"]))

	c.showPage()
	c.save()

def createIMSBoL():
	#Creates the BoL we send to IMS in the same method as the regular BoL
	global master_metadata
	global master_ACE_data
	global config_data

	#Counts how many gaylords are headed to IMS
	fedex_gaylords = []
	fedex_count = 0
	dhl_gaylords = []
	dhl_count = 0
	for gaylord in master_metadata["gaylordAssignments"]:
		if gaylord["carrier"] in config_data["FEDEX_carriers_names"]:
			fedex_gaylords.append(gaylord["id"])
			fedex_count = fedex_count + 1
		elif gaylord["carrier"] in config_data["DHL_carriers_names"]:
			dhl_gaylords.append(gaylord["id"])
			dhl_count = dhl_count + 1
	#Fedex Package Count for emails to Jen @ FedEx
	master_metadata["fedex_package_count"] = 0
	for entry in master_ACE_data:
		if entry["carrier"] == "FEDEX":
			master_metadata["fedex_package_count"] += 1
	
	file_name = master_metadata["date"] + os.sep + master_metadata["date"] + "-IMS-BoL.pdf"
	image_file_path = "resources/IMS_BOL.jpg"
	c = canvas.Canvas(file_name, pagesize = (1668, 1986), bottomup = 0)
	image = ImageReader(image_file_path)
	c.drawImage(image, 0, 0, mask = "auto")

	c.setFont("Courier", 36)
	c.drawString(1070, 184, master_metadata["date"])
	c.drawString(1145, 760, str(fedex_count + dhl_count))
	c.drawString(240, 1463, str(fedex_count) + " pkgs: " + str(master_metadata["fedex_package_count"]))
	c.drawString(705, 1463, str(dhl_count))

	#Writes a list of FedEx gaylords to the sheet in the proper area
	for i, g in enumerate(fedex_gaylords):
		c.drawString(16, 1508 + (i * 36), g)
	#Same for DHL
	for i, g in enumerate(dhl_gaylords):
		c.drawString(520, 1508 + (i * 36), g)

	c.showPage()
	c.save()

def createProForma():
	#Creates the template we upload to SmartBorder
	#Everything is hard-coded to meet their upload schema. Sorry. 
	global master_ACE_data
	global master_metadata
	try:
		if master_ACE_data == []: raise Exception
		else: #If loading didn't error out
			commodities_list = {}
			for entry in master_ACE_data:
				for commodity in entry["commodities"]:
					name = commodity["description"]
					if name not in commodities_list.keys():
						commodities_list[name] = 0 #Add commodity to the list with 0 quantity
					commodities_list[name] = commodities_list[name] + int(commodity["quantity"]) #Update the quantity

			#Techship passes bad data. Error Correction below
			commodities_list = cleanCommoditiesList(commodities_list)

			fda_list = []
			try:
				with open("resources/MASTER_FDA_LIST.csv", "r") as master_file:
					csv_reader = csv.reader(master_file, delimiter = ",")
					for line in csv_reader:
						if line != "":
							fda_list.append(line)
			except: errorBox("Error loading MASTER_FDA_LIST.csv. File open in another program?")

			master_proforma_data = []
			proforma_lines_data = []
			for commodity in commodities_list:
				for line in fda_list:
					if cleanString(commodity).upper() == cleanString(line[2]).upper() and commodities_list[commodity] != 0 and commodity != "" and line[5] != "NOT SHIPPED": #If commodity description matches description from Master FDA file and there is >0 items
						quantity = commodities_list[commodity]
						#Ugly CSV line building
						try:
							#For Master ProForma Upload
							out = (line[1], quantity, "PCS", float(line[5]), "", "", "", line[9], line[10], "", line[12], float(quantity * float(line[5])), "", line[15], line[16], line[17], line[18], line[19], line[20], line[21], line[22], "", line[24], line[25], line[26], line[27], line[28], line[29], line[30], line[31], "", line[33], line[34], line[35], line[36], line[37], line[38], line[39], line[40], "", line[41], float(quantity * float(line[5])), 1, "KG")
							master_proforma_data.append(out)

							#For ProForma Lines Upload
							out = (line[1], quantity, "PCS", float(line[5]), "", "", "", line[9], line[10], "", line[12], float(quantity * float(line[5])), "", line[15], "", "", line[16], line[17], line[18], line[19], line[20], line[21], line[22], "", line[24], "", "", line[25], line[26], line[27], line[28], line[29], line[30], line[31], "", line[32], line[33], "", "", line[34], line[35], line[36], line[37], line[38], line[39], line[40], "", line[41], float(quantity * float(line[5])), 1, "KG")
							proforma_lines_data.append(out)
						except:
							traceback.print_exc()
							print("Unable to output to Proforma:", commodity)
			#More ugly
			master_proforma_header = (("ShipperRefNum","PostToBroker","InvoiceDate","StateDest","PortEntry","MasterBillofLading","Carrier","EstDateTimeArrival","TermsofSale","RelatedParties","ModeTrans","ExportReason","FreightToBorder","ContactInformation","IncludesDuty","FreightAmount","IncludesBrokerage","Currency","TotalGrossWeightKG","ShippingQuantity","ShippingUOM","DutyandFeesBilledTo","InvoiceNumber","OwnerOfGoods","PurchaseOrder","ShipperCustNo","ShipperName","ShipperTaxID","ShipperAddr1","ShipperAddr2","ShipperCity","ShipperState","ShipperCountry","ShipperPostalCode","ShipperMfgID","ShipToCustNo","ShipToName","ShipToTaxID","ShipToAddr1","ShipToAddr2","ShipToCity","ShipToState","ShipToCountry","ShipToPostalCode","SellerCustNo","SellerName","SellerTaxID","SellerAddr1","SellerAddr2","SellerCity","SellerState","SellerCountry","SellerPostalCode","MfgCustNo","MfgName","MfgID","MfgAddress1","MfgAddress2","MfgCity","MfgState","MfgCountry","MfgPostalCode","BuyerCustNo","BuyerName","BuyerUSTaxID","BuyerAddress1","BuyerAddress2","BuyerCity","BuyerState","BuyerCountry","BuyerPostalCode","ConsigneeCustNo","ConsigneeName","ConsigneeUSTaxID","ConsigneeAddress1","ConsigneeAddress2","ConsigneeCity","ConsigneeState","ConsigneeCountry","ConsigneePostalCode","PartNumber","Quantity","QuantityUOM","UnitPrice","GrossWeightKG","NumberOfPackages","PackageUOM","CountryOrigin","SPI","ProductClaimCode","Description","ValueOfGoods","LineMfgCustNo","LineMfgName","LineMfgID","LineMfgAddress1","LineMfgAddress2","LineMfgCity","LineMfgState","LineMfgCountry","LineMfgPostalCode","LineBuyerCustNo","LineBuyerName","LineBuyerUSTaxID","LineBuyerAddress1","LineBuyerAddress2","LineBuyerCity","LineBuyerState","LineBuyerCountry","LineBuyerPostalCode","LineConsigneeCustNo","LineConsigneeName","LineConsigneeUSTaxID","LineConsigneeAddress1","LineConsigneeAddress2","LineConsigneeCity","LineConsigneeState","LineConsigneeCountry","LineConsigneePostalCode","LineNote","Tariff1Number","Tariff1ProductValue","Tariff1Quantity1","Tariff1Quantity1UOM","Tariff1Quantity2","Tariff1Quantity2UOM","Tariff1Quantity3","Tariff1Quantity3UOM","Tariff2Number","Tariff2ProductValue","Tariff2Quantity1","Tariff2Quantity1UOM","Tariff2Quantity2","Tariff2Quantity2UOM","Tariff2Quantity3","Tariff2Quantity3UOM","Tariff3Number","Tariff3ProductValue","Tariff3Quantity1","Tariff3Quantity1UOM","Tariff3Quantity2","Tariff3Quantity2UOM","Tariff3Quantity3","Tariff3Quantity3UOM","Tariff4Number","Tariff4ProductValue","Tariff4Quantity1","Tariff4Quantity1UOM","Tariff4Quantity2","Tariff4Quantity2UOM","Tariff4Quantity3","Tariff4Quantity3UOM","Tariff5Number","Tariff5ProductValue","Tariff5Quantity1","Tariff5Quantity1UOM","Tariff5Quantity2","Tariff5Quantity2UOM","Tariff5Quantity3","Tariff5Quantity3UOM","Tariff6Number","Tariff6ProductValue","Tariff6Quantity1","Tariff6Quantity1UOM","Tariff6Quantity2","Tariff6Quantity2UOM","Tariff6Quantity3","Tariff6Quantity3UOM"))
			master_line_start = (master_metadata["BoL"], "FALSE", master_metadata["date"], "NY", "0901", master_metadata["PAPS"], master_metadata["SCAC"], master_metadata["date"].replace("-", "/") + " 03:00 PM","PLANT","","30","","","","","","","",int(master_metadata["totalWeight"]),master_metadata["packageCount"],"PCS","Buyer","","","","","STALCO INC","160901-55044","401 CLAYSON RD","","NORTH YORK","ON","CA","M9M 2H4","XOSTAINC401NOR","","","","","","","","","","","STALCO INC","160901-55044","401 CLAYSON RD","","NORTH YORK","ON","CA","M9M 2H4","","STALCO INC","XOSTAINC401NOR","401 CLAYSON RD","","NORTH YORK","ON","CA","M9M 2H4","","IMS OF WESTERN NY","16-131314301","2540 WALDEN AVE","SUITE 450","BUFFALO","NY","US","14225","","IMS OF WESTERN NY","16-131314301","2540 WALDEN AVE","SUITE 450","BUFFALO","NY","US","14225")
			
			proforma_lines_header = ("PartNumber","Quantity","QuantityUOM","UnitPrice","GrossWeightKG","NumberOfPackages","PackageUOM","CountryOrigin","SPI","ProductClaimCode","Description","ValueOfGoods","LineMfgCustNo","LineMfgName","LineMfgName2Type","LineMfgName2","LineMfgID","LineMfgAddress1","LineMfgAddress2","LineMfgCity","LineMfgState","LineMfgCountry","LineMfgPostalCode","LineBuyerCustNo","LineBuyerName","LineBuyerName2Type","LineBuyerName2","LineBuyerUSTaxID","LineBuyerAddress1","LineBuyerAddress2","LineBuyerCity","LineBuyerState","LineBuyerCountry","LineBuyerPostalCode","LineConsigneeSameAsBuyer","LineConsigneeCustNo","LineConsigneeName","LineConsigneeName2Type","LineConsigneeName2","LineConsigneeUSTaxID","LineConsigneeAddress1","LineConsigneeAddress2","LineConsigneeCity","LineConsigneeState","LineConsigneeCountry","LineConsigneePostalCode","LineNote","Tariff1Number","Tariff1ProductValue","Tariff1Quantity1","Tariff1Quantity1UOM","Tariff1Quantity2","Tariff1Quantity2UOM","Tariff1Quantity3","Tariff1Quantity3UOM","Tariff2Number","Tariff2ProductValue","Tariff2Quantity1","Tariff2Quantity1UOM","Tariff2Quantity2","Tariff2Quantity2UOM","Tariff2Quantity3","Tariff2Quantity3UOM","Tariff3Number","Tariff3ProductValue","Tariff3Quantity1","Tariff3Quantity1UOM","Tariff3Quantity2","Tariff3Quantity2UOM","Tariff3Quantity3","Tariff3Quantity3UOM","Tariff4Number","Tariff4ProductValue","Tariff4Quantity1","Tariff4Quantity1UOM","Tariff4Quantity2","Tariff4Quantity2UOM","Tariff4Quantity3","Tariff4Quantity3UOM","Tariff5Number","Tariff5ProductValue","Tariff5Quantity1","Tariff5Quantity1UOM","Tariff5Quantity2","Tariff5Quantity2UOM","Tariff5Quantity3","Tariff5Quantity3UOM","Tariff6Number","Tariff6ProductValue","Tariff6Quantity1","Tariff6Quantity1UOM","Tariff6Quantity2","Tariff6Quantity2UOM","Tariff6Quantity3","Tariff6Quantity3UOM")

			#Outputs the excel file for SmartBorder upload
			workbook = pyxl.Workbook()
			filename = master_metadata["date"] + os.sep + master_metadata["date"] + "-ProForma_Template.xlsx"
			worksheet = workbook.active
			worksheet.title = "Sheet1"
			worksheet.append(master_proforma_header)
			for row in master_proforma_data:
				worksheet.append(master_line_start + row)
			workbook.save(filename)

			#Again, but for ProForma Lines
			workbook = pyxl.Workbook()
			filename = master_metadata["date"] + os.sep + master_metadata["date"] + "-ProForma_Lines.xlsx"
			worksheet = workbook.active
			worksheet.title = "Sheet1"
			worksheet.append(proforma_lines_header)
			for row in proforma_lines_data:
				worksheet.append(row)
			workbook.save(filename)

			#Creates the USGR Data file which is used by USTM to make USGRs
			with open(master_metadata["date"] + os.sep + master_metadata["date"] + "-USGR_Data.csv", "w", newline = "") as USGR_file:
				csv_writer = csv.writer(USGR_file)
				for row in master_proforma_header:
					csv_writer.writerow(row)
	except:
		errorBox("Error creating ProForma. Did you create the Master JSON yet?")

def createLoadSheet():
	#Creates a useful but non-essential sheet detailing each gaylord, where its going, if it has FDA products, and how many packages it holds
	global master_metadata
	try:
		#Write to file
		file_name = master_metadata["date"] + os.sep + master_metadata["date"] + "-Load_Sheet.pdf"
		c = canvas.Canvas(file_name, pagesize = (595.27, 841.89), bottomup = 0)
		c.setFont("Courier", 12)
		c.drawString(10, 22, app.getEntry("Date:") + " Load Sheet")
		c.drawString(10, 50, "SKID CARRIER       FDA? PACKAGES")
		for i, row in enumerate(master_metadata["gaylordAssignments"], start = 3):
			c.drawString(10, (22 + i * 14), (row["id"].ljust(5, " ") + row["carrier"].ljust(14, " ") + row["hasFDA"].ljust(5, " ") + str(row["packageCount"])))
		c.showPage()
		c.save()
	except Exception:
		traceback.print_exc()

def emailPaperwork():
	#Check if all files exist
	global master_metadata
	
	folder = master_metadata["date"] + os.sep + master_metadata["date"]
	if os.path.exists(folder + "-ACE.json") and \
	   os.path.exists(folder + "-Detailed_Report.csv") and \
	   os.path.exists(folder + "-IMS-BoL.pdf") and \
	   os.path.exists(folder + "-Load_Sheet.pdf") and \
	   os.path.exists(folder + "-Stalco-BoL.pdf") and \
	   app.getEntry("ProFormaFileEntry") != "" and \
	   app.getEntry("ProFormaFileEntry")[-4:] == ".pdf":
		try:
			#Get files
			files = []
			files.append(folder + "-ACE.json")
			files.append(folder + "-Detailed_Report.csv")
			files.append(folder + "-IMS-BoL.pdf")
			files.append(folder + "-Load_Sheet.pdf")
			files.append(folder + "-Stalco-BoL.pdf")
			files.append(app.getEntry("ProFormaFileEntry"))
			
			#Make email
			global config_data
			username = app.stringBox("Username?", "Please enter your Outlook email address:", parent = None) #Not saved on purpose for security reasons
			password = app.stringBox("Password?", "Please enter the password for {}:".format(username), parent = None)
			
			smtp = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
			smtp.starttls()
			smtp.login(username, password)

			message = MIMEMultipart()
			message_text = str(app.getTextArea("EmailTextArea")) + "\n" + str(master_metadata["gaylordCount"]) + " Gaylords total:\n" + str(master_metadata["USPSCount"]) + " for USPS, " + str(master_metadata["DHLCount"]) + " for DHL, " + str(master_metadata["FEDEXCount"]) + " for FedEx\n" + "Please hit \"REPLY ALL\" when responding to this email trail."
			
			message["From"] = username
			recipients = []
			for recipient in config_data["emailRecipients"]:
				recipients.append(recipient)
			message["To"] = ", ".join(recipients)
			message["Subject"] = "Stalco > Buffalo Run > " + app.getEntry("Date:")
			message.attach(MIMEText(message_text, "plain"))

			#Attach files to email
			for path in files:
				part = MIMEBase('application', "octet-stream")
				with open(path, 'rb') as file:
					part.set_payload(file.read())
				encoders.encode_base64(part)
				part.add_header('Content-Disposition', 'attachment; filename="{}"'.format(Path(path).name))
				message.attach(part)

			smtp.send_message(message)
			del message
			smtp.quit()

			#emailFedex(username, password)
			app.infoBox("Done", "Email successfully sent!")
		except:
			errorBox("Email failed!")
	else:
		errorBox("Not all files required for email are present! Make sure all the boxes are filled")
'''
def emailFedex(username, password):
	global master_metadata
	if master_metadata["fedex_package_count"] != 0:
		try:
			smtp = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
			smtp.starttls()
			smtp.login(username, password)

			message = MIMEMultipart()
			fedex_package_count = str(master_metadata["fedex_package_count"])
			message_text = (f"Morning\n\nWe are dropping off {fedex_package_count} FedEx packages today and would like to request a pickup for the next available business day\n\nThank you!")
			
			message["From"] = username
			global data
			message["To"] = config_data["fedexContact"]
			message["Subject"] = "Stalco > FedEx Pickup Request > " + app.getEntry("Date:")
			message.attach(MIMEText(message_text, "plain"))

			smtp.send_message(message)
			del message
			smtp.quit()
		except:
			errorBox("Error sending FedEx email")
'''
def copyPaperwork():
	#Copies everything to the W: drive
	global master_metadata
	try:
		date = master_metadata["date"]
		src = os.getcwd() + os.sep + date
		year = date.split("-")[0]
		month = date.split("-")[1]
		dst = "W:/Logistics/USPS Customs/USPS Customs Paperwork/IMS Invoices - Mixed Shipments/{year}/{month}/{date}".format(year = year, month = month, date = date)
		if not os.path.exists(dst):
			os.mkdir(dst)
		copyTree(src, dst)
	except:
		errorBox("Copying files failed!")

## Main-Page Support Functions

def loadVariables():
	#CONFIG.json is used to keep track of what the latest BoL/PAPS number is
	try:
		with open("resources/CONFIG.json", "r") as variables_file:
			global config_data
			config_data = json.load(variables_file)
			config_data["date"] = str(date.today())
			app.setEntry("Date:", config_data["date"])
			app.setEntry("BoL #:", config_data["BoL"])
			app.setEntry("PAPS #:", config_data["PAPS"])
			app.setEntry("SCAC:", config_data["SCAC"])
			app.setEntry("USGR Date:", config_data["date"])
			app.setEntry("File Date:", app.getEntry("Date:"))
			app.setTextArea("EmailTextArea", config_data["default_email_message"])
	except:
		errorBox("Error loading /resources/CONFIG.json.\nFile missing/open in another program?")

def increaseVariables():
	#Sets date to today and advances BoL and PAPS number by 1
	global config_data
	config_data["date"] = str(date.today())
	app.setEntry("Date:", config_data["date"])

	config_data["BoL"] = str(int(config_data["BoL"]) + 1).zfill(7)
	app.setEntry("BoL #:", config_data["BoL"])

	config_data["PAPS"] = str(int(config_data["PAPS"]) + 1).zfill(6)
	app.setEntry("PAPS #:", config_data["PAPS"])

def decreaseVariables():
	global config_data
	config_data["date"] = str(date.today())
	app.setEntry("Date:", config_data["date"])

	config_data["BoL"] = str(int(config_data["BoL"]) - 1).zfill(7)
	app.setEntry("BoL #:", config_data["BoL"])

	config_data["PAPS"] = str(int(config_data["PAPS"]) - 1).zfill(6)
	app.setEntry("PAPS #:", config_data["PAPS"])

def saveVariables():
	global config_data
	config_data["date"] = app.getEntry("Date:")
	config_data["BoL"] = app.getEntry("BoL #:")
	config_data["PAPS"] = app.getEntry("PAPS #:")

	with open("resources/CONFIG.json", "w") as variables_file:
		json.dump(config_data, variables_file, indent = 4)

## Page 2 Functions

def loadACEManifest():
	#Loads the ACE so the rest of the features can use the data
	#Also says how many packages are on the ACE as a "green light"
	if app.getEntry("File Date:") == "":
		errorBox("Please enter in the date for this ACE")
	else:
		try:
			global ACE_data
			with open(app.getEntry("ACEManifestFileEntry2"), "r") as ACE_file:
				ACE_data = json.load(ACE_file)
				app.setLabel("ACEStatusLabel", "{} ACE Entries Loaded".format(str(len(ACE_data))))
				SCN_ending = ACE_data[0]["shipmentControlNumber"][-2:]
				app.setLabel("SCNLabel", "SCNs currently end with: {}".format(SCN_ending))
		except FileNotFoundError: errorBox("No file entered")
		except: errorBox("Uploaded file was not an ACE Manfiest")

def removeGaylord():
	#Removes a specific gaylord from the ACE, then prints
	try:
		global ACE_data
		gaylord = app.getEntry('Gaylord (eg. "G1"):')[:2].upper()
		good_entries = []
		bad_entries = []
		try:
			for entry in ACE_data:
				if entry["GAYLORD"] != gaylord:
					good_entries.append(entry)
				else:
					bad_entries.append(entry)
		except:
			errorBox("No enties for Gaylord {}".format(gaylord))
		print(f"{str(len(bad_entries))} entries removed from {gaylord}")
		if len(bad_entries) > 0:
			ACE_data = good_entries
			with open(app.getEntry("ACEManifestFileEntry2"), "w") as ACE_file:
				json.dump(ACE_data, ACE_file, indent = 4)
			with open(app.getEntry("ACEManifestFileEntry2") + "-REMOVED_GAYLORDS", "w") as ACE_file:
				json.dump(bad_entries, ACE_file, indent = 4)
			app.infoBox("Done", "Gaylord {} successfully removed".format(gaylord))
			app.setLabel("ACEStatusLabel", "{} ACE Entries Loaded".format(str(len(ACE_data))))
		else:
			app.infoBox("Done", "No entries found in Gaylord {}".format(gaylord))

	except:
		errorBox("Error removing Gaylord {}. Please check gaylord and if ACE was loaded".format(gaylord))

def removeItems():
	#Removes any packages from the ACE that match the BATCHID/ORDERID as requested
	try:
		global ACE_data
		batches = app.getTextArea('batchesTextArea').replace("\n", ",").split(",")
		good_entries = []
		bad_entries = []
		for entry in ACE_data:
			for line in batches:
				if entry["BATCHID"] != line and entry["ORDERID"] != line:
					if entry not in good_entries:
						good_entries.append(entry)
				else:
					if entry not in bad_entries:
						bad_entries.append(entry)

		ACE_data = good_entries
		with open(app.getEntry("ACEManifestFileEntry2"), "w") as ACE_file:
			json.dump(ACE_data, ACE_file, indent = 4)
		with open(app.getEntry("ACEManifestFileEntry2") + "-REMOVED_ORDERS", "w") as ACE_file:
			json.dump(bad_entries, ACE_file, indent = 4)
		app.infoBox("Done", f"{str(len(bad_entries))} entries removed")
		app.setLabel("ACEStatusLabel", "{} ACE Entries Loaded".format(str(len(ACE_data))))

	except:
		errorBox("Error removing specified batches/orders. Please check batches and if ACE was loaded")

def changeSCNs():
	try:
		new_SCN = app.getEntry("New 2 digits:")[:2]
		global ACE_data
		for entry in ACE_data:
			entry["shipmentControlNumber"] = entry["shipmentControlNumber"][:14] + new_SCN
		with open(app.getEntry("ACEManifestFileEntry2"), "w") as ACE_file:
			json.dump(ACE_data, ACE_file, indent = 4)
			app.setLabel("SCNLabel", "SCNs currently end with: {}".format(new_SCN))
		app.infoBox("Done", "SCN ending changed to {}".format(new_SCN))
	except:
		errorBox("Error changing SCN. Please check new SCN and if ACE was loaded")

def splitACE():
	try:
		global ACE_data
		max_ACE_entries = 9999
		for i in range(len(ACE_data) // max_ACE_entries + 1):
			with open("Split_ACE_Manifest_" + str(i) + ".json", "w") as out_file:
				start = i * max_ACE_entries
				end = (i + 1) * max_ACE_entries - 1
				print(start, end)
				json.dump(ACE_data[start:end], out_file)
		app.infoBox("Done", "ACE successfully split!\nPlease check root folder")
	except:
		errorBox("Unable to split ACE Manifest. Did you remember to load the manifest?")

## Page 3 Functions

def convertJSONToCSV():
	#Process JSON file
	try:
		with open(app.getEntry("JSON")) as json_file:
			data = json.load(json_file)
			filepath = "ACE_Manifest_(CSV).csv"
		data_file = open(filepath, "w", newline="")
		csv_writer = csv.writer(data_file)

		for l in data: #for line in data:
			try:
				_consignee_province = l["consignee"]["address"]["stateProvince"]
				_consignee_postal_code = l["consignee"]["address"]["postalCode"]
			except:
				_consignee_province = ""
			try:
				_shipper_name = l["shipper"]["name"]
				_shipper_address = l["shipper"]["address"]["addressLine"]
				_shipper_country = l["shipper"]["address"]["country"]
				_shipper_city = l["shipper"]["address"]["city"]
				_shipper_province = l["shipper"]["address"]["stateProvince"]
				_shipper_postal_code = l["shipper"]["address"]["postalCode"]
			except:
				_shipper_name = "Stalco Inc."
				_shipper_address = "401 Clayson Road"
				_shipper_country =  "CA"
				_shipper_city = "Toronto"
				_shipper_province = "ON"
				_shipper_postal_code = "M9M2H4"
			try:
				_client = l["client"]
				_carrier = l["carrier"]
				_closeDate = l["closeDate"]
				_trackingNumber = l["trackingNumber"]
				_gaylord = l["GAYLORD"]
			except:
				_client = "N/A"
				_carrier = "N/A"
				_closeDate = "N/A"
				_trackingNumber = "N/A"
				_gaylord = "N/A"
			
			head = ( #Doing it manually for now. This format doesn't change often
				l["ORDERID"],
				l["BATCHID"],
				l["data"],
				l["type"],
				l["shipmentControlNumber"],
				# Defaults for when ACE is missing entries
				#_province_of_loading,
				"ON", #2020-11-26 hardcoding Temporarily
				_shipper_name,
				_shipper_address,
				_shipper_country,
				_shipper_city,
				_shipper_province,
				_shipper_postal_code,
				l["consignee"]["name"],
				l["consignee"]["address"]["addressLine"],
				l["consignee"]["address"]["country"],
				l["consignee"]["address"]["city"],
				_consignee_province,
				_consignee_postal_code,
				_client,
				_carrier,
				_closeDate,
				_trackingNumber,
				_gaylord)
			for i, commodity in enumerate(l["commodities"]): #for commodity in line["commodities"]
				body = (
					l["commodities"][i]["description"],
					l["commodities"][i]["quantity"],
					l["commodities"][i]["packagingUnit"],
					l["commodities"][i]["weight"],
					l["commodities"][i]["weightUnit"])
				if "value" in l["commodities"][i].keys():
					body = body + (l["commodities"][i]["value"],)
				if "countryOfOrigin" in l["commodities"][i]:
					body = body + (l["commodities"][i]["countryOfOrigin"],)
				row = head + body
				csv_writer.writerow(row)

		data_file.close()
		app.infoBox("Done", f"Finished converting JSON.\nOutputting to {filepath}")
	except:
		errorBox("Error converting JSON to CSV.\nPlease see console for more details")

def convertCSVToJSON():
	try:
		with open(app.getEntry("CSV")) as csv_file:
			csv_reader = csv.reader(csv_file, delimiter = ',')
			csv_data = []
			for row in csv_reader:
				csv_data.append(row)
				
			#Makes a list of consignees to add to JSON
			consignees = []
			for row in csv_data:
				if row[4] not in consignees:
					consignees.append(row[4])

			#For each consignee, add each entry to JSON
			out_json = []
			for consignee in consignees:
				entry = {}
				for row in csv_data:
					if consignee == row[4]:
						entry = {
							"ORDERID": row[0], #I don't care if this is hard-coded. I'm the one who made the format of the CSV being read
							"BATCHID": row[1],
							"data": row[2],
							"type": row[3],
							"shipmentControlNumber": row[4],
							"provinceOfLoading": row[5],
							"shipper": {
								"name": row[6],
								"address": {
									"addressLine": row[7],
									"country": row[8],
									"city": row[9],
									"stateProvince": row[10],
									"postalCode": row[11].zfill(5)
								}
							},
							"consignee": {
								"name": row[12],
								"address": {
									"addressLine": row[13],
									"country": row[14],
									"city": row[15],
									"stateProvince": row[16],
									"postalCode": row[17].zfill(5)
								}
							},
							"client": row[18],
							"carrier": row[19],
							"closeDate": row[20],
							"trackingNumber": row[21],
							"GAYLORD": row[22],
							"commodities": []
						}

				for row in csv_data: #Searches for commodities that match consignee
					if consignee == row[4]:
						commodity = {}
						if len(row) == 30: #If the entry has value and countryOfOrigin
							commodity = {
								"description": row[23],
								"quantity": float(row[24]),
								"packagingUnit": row[25],
								"weight": int(row[26]),
								"weightUnit": row[27],
								"value": row[28],
								"countryOfOrigin": row[29]
							}
						elif len(row) == 29: #If it has only value
							commodity = {
								"description": row[23],
								"quantity": float(row[24]),
								"packagingUnit": row[25],
								"weight": int(row[26]),
								"weightUnit": row[27],
								"value": row[28],
							}
						else:
							commodity = {
								"description": row[23],
								"quantity": float(row[24]),
								"packagingUnit": row[25],
								"weight": int(row[26]),
								"weightUnit": row[27]
							}
						entry["commodities"].append(commodity)
				out_json.append(entry)
				filepath = "ACE_Manifest_(JSON).json"
			with open(filepath, "w") as json_file:
				json.dump(out_json, json_file, indent=4)
			app.infoBox("Done", "Done converting CSV to JSON.\nOutputting to {}".format(filepath))
	except:
		errorBox("Error converting from CSV to JSON.\nPlease see console for more details")

def jsonBeautifier():
	#Adds indents to the JSON to make it human-readable
	try:
		json_file_name = app.getEntry("Ugly JSON")
		with open(json_file_name, "r") as json_file:
			json_data = json.load(json_file)
		with open(json_file_name, "w") as json_file:
			json.dump(json_data, json_file, indent = 4)
		app.infoBox("Done", "Done formatting JSON.\nOutputting to {}".format(json_file_name))
	except:
		errorBox("Error beautifying JSON. Make sure uploaded file was valid JSON")
	
def combineJSON():
	#Combines 2 JSONs
	out_data = []
	with open(app.getEntry("JSON 1"), "r") as json_file_1:
		json_data_1 = json.load(json_file_1)
	with open(app.getEntry("JSON 2"), "r") as json_file_2:
		json_data_2 = json.load(json_file_2)
	for line in json_data_1:
		out_data.append(line)
	for line in json_data_2:
		out_data.append(line)
	with open("ACE_Manifest_(Combined).json", "w") as json_file:
		json.dump(out_data, json_file, indent = 4)
	app.infoBox("Done", "Done combining JSONs.\nOutputting to \"ACE_Manifest_(Combined).json\"")

def createUSGR():
	#USGR is paperwork for all the FDA-regulated products that return to the US. For Customs purposes or something
	USGR_date = app.getEntry("USGR Date:").strip()
	USGR_BoL = app.getEntry("USGR BoL #:").strip()
	USGR_entry = app.getEntry("USGR Entry Number:").strip()

	with open("resources/USGR_MASTER_LIST.csv", "r") as USGR_master_file:
		csv_reader = csv.reader(USGR_master_file, delimiter = ',')
		usgr_data = []
		for row in csv_reader:
			usgr_data.append(row)

	with open(app.getEntry("USGR Data:"), "r") as proforma_file:
		csv_reader = csv.reader(proforma_file, delimiter = ',')
		proforma_data = []
		for row in csv_reader:
			proforma_data.append(row)

	#Compares the USGR template data to the USGR Master list. If an item matches and originates from the US, add it to the output data
	out_data = []
	for line in proforma_data:
		for row in usgr_data:
			if line[0] == row[1] and row[10] == "US": #If ProForma SKU matches a SKU from the master file
				out_data.append((row[1], row[13], row[16], row[20], row[21], line[11], line[1], row[2], "BUFFALO", "US"))
	
	#Creates the main USGR information chart
	file_name = "USGR" + os.sep + USGR_date + "-USGR_Table-" + USGR_entry + ".pdf"
	c = canvas.Canvas(file_name, pagesize = (595.27, 841.89), bottomup = 1)
	c.setFont("Courier", 7)
	c.drawString(10, 820, "US GOODS RETURNED")
	c.drawString(10, 812, "INVOICE REFERENCE #: {}    ENTRY #: {}    DATE: {}".format(USGR_BoL, USGR_entry, USGR_date))
	c.drawString(10, 800, "PART/ITEM #       DESCRIPTION                     MANUFACTURER           CITY, STATE           VALUE     QTY  IMPORTDATE  ENTRY PORT")
	for i, row in enumerate(out_data, start = 1):
		out = ""
		out += row[0][:17].ljust(18, " ")
		out += row[1][:30].ljust(32, " ")
		out += row[2][:22].ljust(23, " ")
		out += (row[3][:17] + ", " + row[4][:2]).ljust(22, " ")
		out += row[5][:9].ljust(10, " ")
		out += row[6][:4].ljust(5, " ")
		out += row[7][:11].ljust(12, " ")
		out += "BUFFALO US"
		c.drawString(10, (800 - (i * 8)), str(out))
	c.showPage()

	#Attaches all 6 of the the USGR documents. Adds Entry Number + BoL Number to some of the pages
	c.setFont("Courier", 12)
	for i in range(1, 7): #1 through 6 inclusive
		filepath = "resources" + os.sep + "page_" + str(i) + ".jpg"
		image = Image.open(filepath)
		c.drawImage(ImageReader(image), 0, 0, 595.27, 841.89)
		if i == 1:
			c.drawString(180, 615, app.getEntry("USGR Entry Number:"))
			c.drawString(180, 595, app.getEntry("USGR BoL #:"))
		if i == 2:
			c.drawString(130, 660, app.getEntry("USGR Entry Number:"))
			c.drawString(130, 640, app.getEntry("USGR BoL #:"))
		if i == 3:
			c.drawString(170, 605, app.getEntry("USGR Entry Number:"))
			c.drawString(170, 585, app.getEntry("USGR BoL #:"))
		c.showPage()

	c.save()
	print(f"USGR for {USGR_date} completed")

## Utility Functions

def cleanCommoditiesList(commodities_list: dict):
	#This is kinda dumb
	#CONFIG.json contains an array of key:value pairs
	#loops through this array, matches value to commodities_list, and increments the corresponding key
	global config_data
	commodity_conversion_table = config_data["commodity_conversions"]

	#adds any missing keys to the commodity list
	for commodity in commodity_conversion_table:
		if not commodity in commodities_list:
			commodities_list[commodity] = 0

	#conversion loop
	for commodity in commodities_list:
		for key, value in commodity_conversion_table.items():
			if value == commodity:
				print(value + " converted to " + key)
				commodities_list[key] += commodities_list[value]

	#Kludge bug-fix: go through everything again and set all those converted SKUs' quantities to 0
	for commodity in commodities_list:
		for key, value in commodity_conversion_table.items():
			if value == commodity:
				print(value + " cleared")
				commodities_list[value] = 0

	return commodities_list

def createOutputFolder(folder_string):
	folder_path = os.getcwd() + os.sep + folder_string
	if not os.path.exists(folder_path):
		os.mkdir(folder_path)

def cleanString(in_string):
	return "".join(c for c in in_string if c in "0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ ")

def errorBox(input_string):
	app.errorBox("ErrorBox", input_string)
	try: traceback.print_exc()
	except: pass

def checkFiles():
	return_flag = True
	if app.getEntry("batchesFileEntry") == "":
		errorBox("No Batches Scans file selected")
		return_flag = False
	if app.getEntry("ACEManifestFileEntry") == "":
		errorBox("No ACE Manifest file selected")
		return_flag = False
	if app.getEntry("XLSXReportFileEntry") == "":
		errorBox("No XLSX Report file selected")
		return_flag = False
	if app.getEntry("batchesFileEntry")[-4] != ".csv":
		errorBox("Batches Scans is not .csv")
		return_flag = False
	if app.getEntry("ACEManifestFileEntry")[-5] != ".json":
		errorBox("ACE Manifest is not .json")
		return_flag = False
	if app.getEntry("XLSXReportFileEntry")[-5] != ".xlsx":
		errorBox("XLSX Report is not .csv")
		return_flag = False
	return return_flag
''' Legacy, needs re-write
def updateIMSTracker():
	global master_metadata
	fedex_out = []
	dhl_out = []
	for g in master_metadata["gaylords_assignments"]:
		if g[1] == "DHLGLOBALMAIL":
			dhl_out.append(g[0])
		elif g[1] == "FEDEX":
			fedex_out.append(g[0])

	values = gAPI.main(fedex_in = fedex_out, dhl_in = dhl_out)
	print(values)
'''
### Start

print("USTruckManager.py v2.0")

## NEW Global Variables
consolidated_ACE_data = []
master_ACE_data = []
master_metadata = {}
config_data = {}

## Start GUI
app = gui()

app.startTabbedFrame("TabbedFrame")
app.setSticky("nesw")
app.setStretch("column")

## Frame 1
app.startTab("BASIC")

app.startLabelFrame("Date/BoL/PAPS")
app.addLabel("W:\\Logistics\\Carrier Tracking\\USPS Tracking.xlsx")
app.addLabelEntry("Date:")
app.addLabelEntry("BoL #:")
app.addLabelEntry("PAPS #:")
app.addLabelEntry("SCAC:")
app.addButton("+1 BoL #/PAPS #", increaseVariables)
app.addButton("-1 BoL #/PAPS #", decreaseVariables)
app.addButton("Save BoL #/PAPS #", saveVariables)
app.stopLabelFrame()

app.startLabelFrame("Step 1")
app.addLabel("ACE Manfiest (from Techship, \"ace_manifest_#\"):")
app.addFileEntry("ACEManifestFileEntry")
app.addLabel("Detailed Report:")
app.addFileEntry("batchesFileEntry")
app.addLabel("Report XLSX (from Techship, \"manifest_packages_#\"):")
app.addFileEntry("XLSXReportFileEntry")
app.addButton("Create Paperwork", doEverything)
app.stopLabelFrame()

app.startLabelFrame("Step 2:")
app.addLabel("ProForma (Printed from SmartBorder):")
app.addFileEntry("ProFormaFileEntry")
app.setStretch("both")
app.setSticky("nesw")
app.addTextArea("EmailTextArea", text = None)
app.setSticky("ws")
app.setStretch("column")
app.addButton("Email Paperwork", emailPaperwork)
#app.addButton("Update IMS Tracker", updateIMSTracker)
app.addButton("Move Paperwork to W: Drive", copyPaperwork)
app.stopLabelFrame()
app.stopTab()

## Frame 2
app.startTab("ACE EDITING")

app.startLabelFrame("ACE Manifest")
app.addLabel("ACE Manifest (.json):")
app.addFileEntry("ACEManifestFileEntry2")
app.addLabelEntry("File Date:")
app.addButton("Load ACE", loadACEManifest)
app.addLabel("ACEStatusLabel", "No ACE Loaded")
app.stopLabelFrame()

app.startLabelFrame("Gaylord Removal")
app.addLabelEntry('Gaylord (eg. "G1"):')
app.addButton("Remove Gaylord", removeGaylord)
app.stopLabelFrame()

app.startLabelFrame("Batch/Order Removal")
app.addLabel('Batches/Orders:')
app.addTextArea("batchesTextArea")
app.addButton("Remove Items", removeItems)
app.stopLabelFrame()

app.startLabelFrame("Duplicate SCN Editor")
app.addLabel("SCNLabel", "SCNs currently end with: NA")
app.addLabelEntry("New 2 digits:")
app.addButton("Change SCNs", changeSCNs)
app.stopLabelFrame()

app.startLabelFrame("ACE Splitter")
app.addLabel("Use if ACE exceeds 9999 entries:")
app.addButton("Split", splitACE)
app.stopLabelFrame()

app.stopTab()

## Frame 3
app.startTab("ADVANCED")

app.startLabelFrame("Convert ACE to EXCEL:")
app.addLabelFileEntry("JSON")
app.addButton("Convert to CSV", convertJSONToCSV)
app.addLabelFileEntry("CSV")
app.addButton("Convert to JSON", convertCSVToJSON)
app.stopLabelFrame()

app.startLabelFrame("JSON Formatter")
app.addLabelFileEntry("Ugly JSON")
app.addButton("Format JSON", jsonBeautifier)
app.stopLabelFrame()

app.startLabelFrame("JSON Combiner")
app.addLabelFileEntry("JSON 1")
app.addLabelFileEntry("JSON 2")
app.addButton("Combine", combineJSON)
app.stopLabelFrame()

app.stopTab()

## Frame 4
app.startTab("USGR")
app.startLabelFrame("USGR")
app.addLabelFileEntry("USGR Data:")
app.addLabelEntry("USGR Date:")
app.addLabelEntry("USGR Entry Number:")
app.addLabelEntry("USGR BoL #:")
app.addButton("Create USGR", createUSGR)
app.stopLabelFrame()
app.stopTab()

## Finish GUI
app.stopTabbedFrame()

### Setup

loadVariables()

app.go()
